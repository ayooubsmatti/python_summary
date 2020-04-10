# =============================================================
# import openpyxl as xl
# wb = xl.load_workbook('transactions.xlsx')
# sheet = wb['Sheet1']
# cell = sheet['a1']
# for row in range(2, sheet.max_row + 1):
#     cell = sheet.cell(row, 3)
#     correct_price = cell.value * 0.9
#     correct_price_cell = sheet.cell(row, 4)
#     correct_price_cell.value = correct_price
#
# wb.save('transactions2.xlsx')
# ==============================================================
# Bar chart
import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_worbook(filname,namefile):
        wb = xl.load_workbook(filname)
        sheet = wb['Sheet1']
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, 3)
            correct_price = cell.value * 0.9
            correct_price_cell = sheet.cell(row, 4)
            correct_price_cell.value = correct_price

        values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
        chart = BarChart()
        chart.add_data(values)
        sheet.add_chart(chart, 'd2')

        wb.save(namefile)


process_worbook("transactions.xlsx", "transactions3.xlsx")
